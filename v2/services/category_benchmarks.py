"""Category benchmarks — parsing ML's XLSX exports for competitive analytics.

ML provides three relevant XLSX exports manually (no API equivalent):

1. **benchmark_categories-Relatorio_composicao_vendas_*.xlsx**
   Sheets: "Vendas por categorias", "Vendas por subcategorias".
   Per-seller distribution of own sales by category/subcategory.

2. **Mais_vendidos_nas_suas_categorias_*.xlsx**
   Sheets per category (Calçados/Roupas/Bolsas, Malas e Bolsas, ...).
   Top-100 listings in each category WITH metrics (visits, sales, conversion,
   photo count, video, ads, parcelas, frete). This is the goldmine for
   competitive benchmarks — ML's only public source of competitor visits.

3. **Relatorio_desempenho_publicacoes_*.xlsx**
   Per-listing performance for the seller's own items (visits, sales, reviews
   total/good/bad, "Experiência de compra" composite label). Adds reviews
   and experiência data that we don't have via API.

Tables:
  category_sales_composition (user, period, category, subcat, brl, pct, ...)
  category_top_listings      (user, subcat, period, name, image, price, ...)
  category_benchmarks        (user, subcat, median_price, avg_photos, ...)
  listing_performance_snapshots (user, item_id, variation, reviews, exp, ...)

Pattern: TEST → DB → CACHE — XLSX upload IS the test+refresh; benchmark
aggregation is the cache layer that downstream UI consumes.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Any, Optional

import asyncpg
import openpyxl

log = logging.getLogger(__name__)


CREATE_SQL = """
-- File 1: composição (own sales distribution)
CREATE TABLE IF NOT EXISTS category_sales_composition (
  id SERIAL PRIMARY KEY,
  user_id INTEGER NOT NULL,
  period_start DATE,
  period_end DATE,
  category TEXT,
  subcategory TEXT,
  gross_brl NUMERIC,
  pct_participacao NUMERIC,
  buyers INTEGER,
  conversion NUMERIC,
  imported_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_csc_user ON category_sales_composition(user_id, period_end DESC);

-- File 2: top-100 in each subcategory (raw)
CREATE TABLE IF NOT EXISTS category_top_listings (
  id SERIAL PRIMARY KEY,
  user_id INTEGER NOT NULL,
  subcategory TEXT,
  period_start DATE,
  period_end DATE,
  rank INTEGER,
  name TEXT,
  image_url TEXT,
  price NUMERIC,
  units_sold INTEGER,
  views INTEGER,
  cancelled INTEGER,
  questions INTEGER,
  photos INTEGER,
  has_video BOOLEAN,
  has_ads BOOLEAN,
  has_parcelas BOOLEAN,
  is_catalog BOOLEAN,
  frete_type TEXT,
  imported_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_ctl_user_subcat ON category_top_listings(user_id, subcategory);

-- Computed aggregates per subcategory
CREATE TABLE IF NOT EXISTS category_benchmarks (
  user_id INTEGER NOT NULL,
  subcategory TEXT NOT NULL,
  period_start DATE,
  period_end DATE,
  sample_size INTEGER,
  median_price NUMERIC,
  avg_photos NUMERIC,
  pct_full NUMERIC,
  pct_video NUMERIC,
  pct_ads NUMERIC,
  pct_parcelas NUMERIC,
  pct_catalog NUMERIC,
  median_conversion NUMERIC,
  median_questions NUMERIC,
  median_views INTEGER,
  median_units_sold INTEGER,
  computed_at TIMESTAMPTZ DEFAULT NOW(),
  PRIMARY KEY (user_id, subcategory)
);

-- File 3: per-listing performance (own items, reviews + experiência)
CREATE TABLE IF NOT EXISTS listing_performance_snapshots (
  id SERIAL PRIMARY KEY,
  user_id INTEGER NOT NULL,
  item_id TEXT,
  variation TEXT,
  sku TEXT,
  status TEXT,
  quality TEXT,                -- Básica / Profissional
  experiencia TEXT,            -- Boa / Média / Ruim
  unique_visits INTEGER,
  sales_count INTEGER,
  unique_buyers INTEGER,
  units_sold INTEGER,
  gross_brl NUMERIC,
  participacao NUMERIC,
  conv_visits_sales NUMERIC,
  conv_visits_buyers NUMERIC,
  reviews_total INTEGER,
  reviews_bad INTEGER,
  reviews_good INTEGER,
  period_start DATE,
  period_end DATE,
  imported_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_lps_user_item ON listing_performance_snapshots(user_id, item_id);
"""


async def ensure_schema(pool: asyncpg.Pool) -> None:
    async with pool.acquire() as conn:
        await conn.execute(CREATE_SQL)


# ── Type detection ────────────────────────────────────────────────────────────

class XlsxType:
    COMPOSICAO = "composicao_vendas"          # File 1
    MAIS_VENDIDOS = "mais_vendidos_categoria"  # File 2
    DESEMPENHO = "desempenho_publicacoes"      # File 3
    UNKNOWN = "unknown"


def detect_xlsx_type(filename: str, sheet_names: list[str]) -> str:
    """Detect XLSX type from filename + sheet structure."""
    lower = (filename or "").lower()
    sheets_lower = [s.lower() for s in sheet_names]

    if "benchmark_categor" in lower or "composicao_vendas" in lower or "composição" in lower:
        return XlsxType.COMPOSICAO
    if "vendas por categoria" in " ".join(sheets_lower) or "vendas por subcategoria" in " ".join(sheets_lower):
        return XlsxType.COMPOSICAO

    if "mais_vendidos" in lower or "mais vendidos" in lower:
        return XlsxType.MAIS_VENDIDOS

    if "desempenho" in lower or "publicacoes" in lower:
        return XlsxType.DESEMPENHO
    if any("relatório" in s for s in sheets_lower) and any("definições" in s for s in sheets_lower):
        # Generic ML format — fall back on column sniff
        return XlsxType.UNKNOWN

    return XlsxType.UNKNOWN


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_brl(v: Any) -> Optional[float]:
    """ML's BRL cells come as 'R$ 2.737' or '2.737,65'. Strip and parse."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("R$", "").replace("\xa0", " ").strip()
    if not s:
        return None
    # Brazilian format: dot=thousands, comma=decimal. Remove dots first.
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_pct(v: Any) -> Optional[float]:
    """'30,1%' → 30.1; '0,7%' → 0.7."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) * 100.0 if abs(float(v)) <= 1 else float(v)
    s = str(v).replace("%", "").strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(".", "").replace(" ", "").strip()
    if not s or s == "-":
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_sim_nao(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("sim", "yes", "true", "1"):
        return True
    if s in ("não", "nao", "no", "false", "0"):
        return False
    return None


def _extract_period_from_filename(filename: str) -> tuple[Optional[date], Optional[date]]:
    """Extract YYYY-MM-DD..YYYY-MM-DD from typical ML filenames."""
    matches = re.findall(r"(\d{4})[_-](\d{2})[_-](\d{2})", filename or "")
    if len(matches) >= 2:
        try:
            start = date(int(matches[0][0]), int(matches[0][1]), int(matches[0][2]))
            end = date(int(matches[1][0]), int(matches[1][1]), int(matches[1][2]))
            return start, end
        except (ValueError, TypeError):
            pass
    return None, None


# ── Parser: Composição ────────────────────────────────────────────────────────

def parse_composicao(
    file_bytes: bytes, filename: str = "",
) -> list[dict[str, Any]]:
    """File 1: own sales by category + subcategory."""
    period_start, period_end = _extract_period_from_filename(filename)
    rows: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # Subcategories sheet has more granular data with conversion
    target = None
    for s in wb.sheetnames:
        if "subcategor" in s.lower():
            target = s
            break
    if not target:
        for s in wb.sheetnames:
            if "categoria" in s.lower():
                target = s
                break
    if not target:
        return rows

    ws = wb[target]
    # Header row 5 (1-indexed). Data starts row 6.
    header_row = None
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if r and r[0] and "categoria" in str(r[0]).lower():
            header_row = i
            break
    if header_row is None:
        header_row = 5

    has_subcat = "subcategor" in target.lower()
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or not r[0]:
            continue
        cat = str(r[0]).strip()
        if has_subcat:
            subcat = str(r[1]).strip() if r[1] else None
            brl = _parse_brl(r[2])
            pct_buyers = _parse_pct(r[3]) if len(r) > 3 else None
            buyers = _parse_int(r[4]) if len(r) > 4 else None
            pct_part = _parse_pct(r[5]) if len(r) > 5 else None
            conv = _parse_pct(r[6]) if len(r) > 6 else None
        else:
            subcat = None
            brl = _parse_brl(r[1])
            pct_part = _parse_pct(r[2]) if len(r) > 2 else None
            buyers = None
            conv = None
        if not cat:
            continue
        rows.append({
            "category": cat, "subcategory": subcat,
            "gross_brl": brl, "pct_participacao": pct_part,
            "buyers": buyers, "conversion": conv,
            "period_start": period_start, "period_end": period_end,
        })
    return rows


# ── Parser: Mais Vendidos ─────────────────────────────────────────────────────

# Image URL pattern: D_<picid>-MLB<itemid>_<date>-<variant>.<ext>
_IMG_MLB = re.compile(r"-MLB(\d+)_")


def parse_mais_vendidos(
    file_bytes: bytes, filename: str = "",
) -> list[dict[str, Any]]:
    """File 2: top-100 listings per category (sheet = subcategory)."""
    period_start, period_end = _extract_period_from_filename(filename)
    rows: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name in ("Definições",) or sheet_name.lower().startswith("nombrecategoria"):
            continue
        ws = wb[sheet_name]
        # Header row 6, data row 7+
        rank = 0
        for r_idx, r in enumerate(ws.iter_rows(min_row=7, values_only=True), 7):
            if not r or len(r) < 4:
                continue
            # Skip placeholder rows (template "{CAT_L3_TITLE}" etc)
            if r[0] is None and r[1] is None:
                continue
            name = str(r[1]).strip() if r[1] else None
            if not name or name == "-":
                continue
            rank += 1
            image_url = str(r[0]).strip() if r[0] else None
            condition = str(r[2]).strip() if r[2] else None
            price = _parse_brl(r[3])
            units = _parse_int(r[4])
            views = _parse_int(r[5])
            cancelled = _parse_int(r[6])
            questions = _parse_int(r[7])
            is_cat = _parse_sim_nao(r[8]) if len(r) > 8 else None
            has_video = _parse_sim_nao(r[9]) if len(r) > 9 else None
            has_parc = _parse_sim_nao(r[10]) if len(r) > 10 else None
            has_ads = _parse_sim_nao(r[11]) if len(r) > 11 else None
            photos = _parse_int(r[12]) if len(r) > 12 else None
            frete = str(r[13]).strip() if len(r) > 13 and r[13] else None

            rows.append({
                "subcategory": sheet_name,
                "rank": rank,
                "name": name,
                "image_url": image_url,
                "price": price,
                "units_sold": units,
                "views": views,
                "cancelled": cancelled,
                "questions": questions,
                "photos": photos,
                "has_video": has_video,
                "has_ads": has_ads,
                "has_parcelas": has_parc,
                "is_catalog": is_cat,
                "frete_type": frete,
                "period_start": period_start,
                "period_end": period_end,
            })
    return rows


# ── Parser: Desempenho de Publicações ─────────────────────────────────────────

def parse_desempenho(
    file_bytes: bytes, filename: str = "",
) -> list[dict[str, Any]]:
    """File 3: per-listing performance (own items)."""
    period_start, period_end = _extract_period_from_filename(filename)
    rows: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    target = None
    for s in wb.sheetnames:
        if "relat" in s.lower() or "publicac" in s.lower():
            target = s
            break
    if not target:
        return rows

    ws = wb[target]
    # Header row 6, data row 7+. Cols (0-indexed):
    # 0:item_id, 1:title, 2:status, 3:variation, 4:sku, 5:quality, 6:experiencia,
    # 7:unique_visits, 8:sales_count, 9:unique_buyers, 10:units_sold,
    # 11:gross_brl, 12:participacao, 13:conv_v_s, 14:conv_v_b,
    # 15:reviews_total, 16:reviews_bad, 17:reviews_good
    for r in ws.iter_rows(min_row=7, values_only=True):
        if not r or not r[0]:
            continue
        item_id = str(r[0]).strip()
        if not item_id.isdigit() and not item_id.startswith("MLB"):
            continue
        if item_id.isdigit():
            item_id = "MLB" + item_id
        rows.append({
            "item_id": item_id,
            "variation": str(r[3]).strip() if len(r) > 3 and r[3] else None,
            "sku": str(r[4]).strip() if len(r) > 4 and r[4] else None,
            "status": str(r[2]).strip() if len(r) > 2 and r[2] else None,
            "quality": str(r[5]).strip() if len(r) > 5 and r[5] else None,
            "experiencia": str(r[6]).strip() if len(r) > 6 and r[6] else None,
            "unique_visits": _parse_int(r[7]) if len(r) > 7 else None,
            "sales_count": _parse_int(r[8]) if len(r) > 8 else None,
            "unique_buyers": _parse_int(r[9]) if len(r) > 9 else None,
            "units_sold": _parse_int(r[10]) if len(r) > 10 else None,
            "gross_brl": _parse_brl(r[11]) if len(r) > 11 else None,
            "participacao": _parse_pct(r[12]) if len(r) > 12 else None,
            "conv_visits_sales": _parse_pct(r[13]) if len(r) > 13 else None,
            "conv_visits_buyers": _parse_pct(r[14]) if len(r) > 14 else None,
            "reviews_total": _parse_int(r[15]) if len(r) > 15 else None,
            "reviews_bad": _parse_int(r[16]) if len(r) > 16 else None,
            "reviews_good": _parse_int(r[17]) if len(r) > 17 else None,
            "period_start": period_start,
            "period_end": period_end,
        })
    return rows


# ── DB upserts ────────────────────────────────────────────────────────────────

async def store_composicao(pool: asyncpg.Pool, user_id: int, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    saved = 0
    async with pool.acquire() as conn:
        # Replace prior rows for the same period to avoid duplicates on re-upload
        period_end = rows[0].get("period_end")
        if period_end:
            await conn.execute(
                "DELETE FROM category_sales_composition WHERE user_id = $1 AND period_end = $2",
                user_id, period_end,
            )
        for r in rows:
            await conn.execute(
                """
                INSERT INTO category_sales_composition
                  (user_id, period_start, period_end, category, subcategory,
                   gross_brl, pct_participacao, buyers, conversion)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                """,
                user_id, r.get("period_start"), r.get("period_end"),
                r.get("category"), r.get("subcategory"),
                r.get("gross_brl"), r.get("pct_participacao"),
                r.get("buyers"), r.get("conversion"),
            )
            saved += 1
    return saved


async def store_top_listings(pool: asyncpg.Pool, user_id: int, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    saved = 0
    async with pool.acquire() as conn:
        # Replace prior period — top-100 is a snapshot, latest wins
        period_end = rows[0].get("period_end")
        if period_end:
            await conn.execute(
                "DELETE FROM category_top_listings WHERE user_id = $1 AND period_end = $2",
                user_id, period_end,
            )
        for r in rows:
            await conn.execute(
                """
                INSERT INTO category_top_listings
                  (user_id, subcategory, period_start, period_end, rank, name, image_url,
                   price, units_sold, views, cancelled, questions, photos,
                   has_video, has_ads, has_parcelas, is_catalog, frete_type)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13,
                        $14, $15, $16, $17, $18)
                """,
                user_id, r.get("subcategory"), r.get("period_start"), r.get("period_end"),
                r.get("rank"), r.get("name"), r.get("image_url"),
                r.get("price"), r.get("units_sold"), r.get("views"),
                r.get("cancelled"), r.get("questions"), r.get("photos"),
                r.get("has_video"), r.get("has_ads"), r.get("has_parcelas"),
                r.get("is_catalog"), r.get("frete_type"),
            )
            saved += 1
    return saved


async def store_desempenho(pool: asyncpg.Pool, user_id: int, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    saved = 0
    async with pool.acquire() as conn:
        period_end = rows[0].get("period_end")
        if period_end:
            await conn.execute(
                "DELETE FROM listing_performance_snapshots WHERE user_id = $1 AND period_end = $2",
                user_id, period_end,
            )
        for r in rows:
            await conn.execute(
                """
                INSERT INTO listing_performance_snapshots
                  (user_id, item_id, variation, sku, status, quality, experiencia,
                   unique_visits, sales_count, unique_buyers, units_sold,
                   gross_brl, participacao, conv_visits_sales, conv_visits_buyers,
                   reviews_total, reviews_bad, reviews_good,
                   period_start, period_end)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11,
                        $12, $13, $14, $15, $16, $17, $18, $19, $20)
                """,
                user_id, r.get("item_id"), r.get("variation"), r.get("sku"),
                r.get("status"), r.get("quality"), r.get("experiencia"),
                r.get("unique_visits"), r.get("sales_count"), r.get("unique_buyers"),
                r.get("units_sold"), r.get("gross_brl"), r.get("participacao"),
                r.get("conv_visits_sales"), r.get("conv_visits_buyers"),
                r.get("reviews_total"), r.get("reviews_bad"), r.get("reviews_good"),
                r.get("period_start"), r.get("period_end"),
            )
            saved += 1
    return saved


# ── Benchmark aggregator ──────────────────────────────────────────────────────

async def compute_benchmarks(pool: asyncpg.Pool, user_id: int) -> int:
    """Re-aggregate category_top_listings → category_benchmarks per subcat.
    Replaces existing rows for the user."""
    async with pool.acquire() as conn:
        await conn.execute(
            "DELETE FROM category_benchmarks WHERE user_id = $1",
            user_id,
        )
        # Aggregate stats per subcategory
        await conn.execute(
            """
            INSERT INTO category_benchmarks
              (user_id, subcategory, period_start, period_end, sample_size,
               median_price, avg_photos, pct_full, pct_video, pct_ads,
               pct_parcelas, pct_catalog,
               median_conversion, median_questions, median_views, median_units_sold,
               computed_at)
            SELECT
              user_id,
              subcategory,
              MIN(period_start),
              MAX(period_end),
              COUNT(*) AS sample_size,
              PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price) AS median_price,
              AVG(photos)::numeric AS avg_photos,
              AVG(CASE WHEN frete_type ILIKE '%full%' THEN 1.0 ELSE 0.0 END) * 100 AS pct_full,
              AVG(CASE WHEN has_video THEN 1.0 ELSE 0.0 END) * 100 AS pct_video,
              AVG(CASE WHEN has_ads THEN 1.0 ELSE 0.0 END) * 100 AS pct_ads,
              AVG(CASE WHEN has_parcelas THEN 1.0 ELSE 0.0 END) * 100 AS pct_parcelas,
              AVG(CASE WHEN is_catalog THEN 1.0 ELSE 0.0 END) * 100 AS pct_catalog,
              PERCENTILE_CONT(0.5) WITHIN GROUP (
                ORDER BY CASE WHEN views > 0 AND units_sold IS NOT NULL
                              THEN units_sold::numeric / views * 100 ELSE NULL END
              ) AS median_conversion,
              PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY questions) AS median_questions,
              PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY views)::int AS median_views,
              PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY units_sold)::int AS median_units_sold,
              NOW()
            FROM category_top_listings
            WHERE user_id = $1 AND subcategory IS NOT NULL
            GROUP BY user_id, subcategory
            """,
            user_id,
        )
        result = await conn.fetchval(
            "SELECT COUNT(*) FROM category_benchmarks WHERE user_id = $1",
            user_id,
        )
    return int(result or 0)


# ── Read APIs ─────────────────────────────────────────────────────────────────

async def get_benchmarks(pool: asyncpg.Pool, user_id: int) -> list[dict[str, Any]]:
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT subcategory, sample_size, median_price, avg_photos,
                   pct_full, pct_video, pct_ads, pct_parcelas, pct_catalog,
                   median_conversion, median_questions, median_views, median_units_sold,
                   to_char(period_end, 'YYYY-MM-DD') AS period_end,
                   to_char(computed_at AT TIME ZONE 'UTC',
                           'YYYY-MM-DD"T"HH24:MI:SS"Z"') AS computed_at
              FROM category_benchmarks
             WHERE user_id = $1
             ORDER BY sample_size DESC NULLS LAST
            """,
            user_id,
        )
    return [
        {
            "subcategory": r["subcategory"],
            "sampleSize": int(r["sample_size"] or 0),
            "medianPrice": float(r["median_price"]) if r["median_price"] is not None else None,
            "avgPhotos": float(r["avg_photos"]) if r["avg_photos"] is not None else None,
            "pctFull": float(r["pct_full"]) if r["pct_full"] is not None else None,
            "pctVideo": float(r["pct_video"]) if r["pct_video"] is not None else None,
            "pctAds": float(r["pct_ads"]) if r["pct_ads"] is not None else None,
            "pctParcelas": float(r["pct_parcelas"]) if r["pct_parcelas"] is not None else None,
            "pctCatalog": float(r["pct_catalog"]) if r["pct_catalog"] is not None else None,
            "medianConversion": float(r["median_conversion"]) if r["median_conversion"] is not None else None,
            "medianQuestions": float(r["median_questions"]) if r["median_questions"] is not None else None,
            "medianViews": int(r["median_views"]) if r["median_views"] is not None else None,
            "medianUnitsSold": int(r["median_units_sold"]) if r["median_units_sold"] is not None else None,
            "periodEnd": r["period_end"],
            "computedAt": r["computed_at"],
        }
        for r in rows
    ]


async def get_top_listings(
    pool: asyncpg.Pool, user_id: int, subcategory: str, limit: int = 100,
) -> list[dict[str, Any]]:
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT rank, name, image_url, price, units_sold, views,
                   cancelled, questions, photos, has_video, has_ads,
                   has_parcelas, is_catalog, frete_type
              FROM category_top_listings
             WHERE user_id = $1 AND subcategory = $2
             ORDER BY rank ASC
             LIMIT $3
            """,
            user_id, subcategory, limit,
        )
    return [
        {
            "rank": int(r["rank"] or 0),
            "name": r["name"],
            "imageUrl": r["image_url"],
            "price": float(r["price"]) if r["price"] is not None else None,
            "unitsSold": r["units_sold"],
            "views": r["views"],
            "cancelled": r["cancelled"],
            "questions": r["questions"],
            "photos": r["photos"],
            "hasVideo": r["has_video"],
            "hasAds": r["has_ads"],
            "hasParcelas": r["has_parcelas"],
            "isCatalog": r["is_catalog"],
            "freteType": r["frete_type"],
        }
        for r in rows
    ]


async def get_listing_performance(
    pool: asyncpg.Pool, user_id: int, item_id: Optional[str] = None,
) -> list[dict[str, Any]]:
    """Per-item performance snapshots — used to enrich item cards with reviews
    and 'Experiência de compra' that ML API doesn't expose."""
    where = "WHERE user_id = $1"
    params: list[Any] = [user_id]
    if item_id:
        where += " AND item_id = $2"
        params.append(item_id)
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"""
            SELECT item_id, variation, sku, status, quality, experiencia,
                   unique_visits, sales_count, units_sold, gross_brl,
                   conv_visits_sales, reviews_total, reviews_bad, reviews_good,
                   to_char(period_end, 'YYYY-MM-DD') AS period_end
              FROM listing_performance_snapshots
              {where}
             ORDER BY period_end DESC NULLS LAST, gross_brl DESC NULLS LAST
            """,
            *params,
        )
    return [
        {
            "itemId": r["item_id"],
            "variation": r["variation"],
            "sku": r["sku"],
            "status": r["status"],
            "quality": r["quality"],
            "experiencia": r["experiencia"],
            "uniqueVisits": r["unique_visits"],
            "salesCount": r["sales_count"],
            "unitsSold": r["units_sold"],
            "grossBrl": float(r["gross_brl"]) if r["gross_brl"] is not None else None,
            "convVisitsSales": float(r["conv_visits_sales"]) if r["conv_visits_sales"] is not None else None,
            "reviewsTotal": r["reviews_total"],
            "reviewsBad": r["reviews_bad"],
            "reviewsGood": r["reviews_good"],
            "periodEnd": r["period_end"],
        }
        for r in rows
    ]
