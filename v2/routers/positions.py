"""Position-in-search endpoints for the Escalar Marketing page.

Calls ML's public `/sites/<site>/search` (no auth) via `services.positions`
and records every check into `position_history` for the authenticated user.
"""
from __future__ import annotations

from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile

from v2.deps import CurrentUser, current_user, get_pool
from v2.schemas.escalar import (
    PositionCheckOut,
    TrackedKeyword,
    TrackedListOut,
    TrackKeywordIn,
)
from v2.services import ml_oauth as ml_oauth_svc
from v2.services import positions as positions_service
from v2.storage import positions_storage

router = APIRouter(prefix="/escalar/positions", tags=["escalar", "positions"])


def _iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None


@router.get("/_probe")
async def positions_probe(
    keyword: str = Query(..., min_length=1),
    site_id: str = Query("MLB"),
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    """Diagnostic: hit ML JSON search API directly with the user's bearer
    + show storage_state status. Used to figure out why /check returns
    ml_no_results — is it the JSON API blocked, scraper blocked, or
    actually empty results from ML?
    """
    if pool is None:
        return {"error": "no_db"}

    import httpx
    from urllib.parse import quote_plus
    from v2.services import ml_oauth as _oauth_svc, ml_scraper as _scraper

    # 1. ML token state
    try:
        token, exp, refreshed = await _oauth_svc.get_valid_access_token(pool, user.id)
        token_state = {
            "ok": True,
            "expires_at": exp.isoformat() if exp else None,
            "refreshed_now": refreshed,
            "token_prefix": token[:12] + "..." if token else None,
        }
    except Exception as err:  # noqa: BLE001
        token = None
        token_state = {"ok": False, "error": str(err)}

    # 2. Storage state (logged-in scraper) — show whether env is set + decoded
    storage_state_info: dict[str, Any] = {"loaded": False}
    state = _scraper._storage_state()
    if state is not None:
        cookies = state.get("cookies") or []
        storage_state_info = {
            "loaded": True,
            "cookies_count": len(cookies),
            "auth_critical_present": [
                c.get("name") for c in cookies
                if c.get("name") in (
                    "ssid", "nsa_rotok", "NSESSIONID_pampa_session",
                    "_csrf", "cp", "ftid", "orguserid",
                )
            ],
            "domains": sorted(set((c.get("domain") or "").lstrip(".") for c in cookies)),
        }

    # 3. JSON API direct call
    json_api: dict[str, Any] = {"called": False}
    if token:
        url = (
            f"https://api.mercadolibre.com/sites/{site_id}/search"
            f"?q={quote_plus(keyword)}&limit=10&offset=0"
        )
        try:
            async with httpx.AsyncClient() as http:
                r = await http.get(
                    url,
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=15.0,
                )
            try:
                body = r.json()
            except Exception:  # noqa: BLE001
                body = None
            sample_ids = []
            paging = None
            if isinstance(body, dict):
                paging = body.get("paging")
                results = body.get("results") or []
                for it in results[:5]:
                    sample_ids.append({
                        "id": it.get("id"),
                        "title": (it.get("title") or "")[:80],
                        "tags": it.get("tags") or [],
                    })
            json_api = {
                "called": True,
                "url": url,
                "status": r.status_code,
                "content_type": r.headers.get("content-type"),
                "body_preview": r.text[:500],
                "paging": paging,
                "sample_first_5": sample_ids,
            }
        except Exception as err:  # noqa: BLE001
            json_api = {"called": True, "url": url, "error": str(err)}

    return {
        "user_id": user.id,
        "keyword": keyword,
        "site_id": site_id,
        "token": token_state,
        "storage_state": storage_state_info,
        "json_api": json_api,
    }


@router.get("/check", response_model=PositionCheckOut)
async def check(
    item_id: str = Query(..., min_length=3),
    keyword: str = Query(..., min_length=1),
    site_id: str = Query("MLB"),
    category_id: Optional[str] = Query(None),
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    """One-shot check: search ML for `keyword`, scan up to 1000 results.

    Reads the user's ML bearer from Railway Postgres `ml_user_tokens`
    (`ml_oauth.get_valid_access_token` auto-refreshes if expiring). No
    Supabase path. If no token, returns 428 → UI deep-links to /escalar/settings.
    """
    if pool is None:
        raise HTTPException(status_code=503, detail="no_db")
    try:
        bearer, _expires, _refreshed = await ml_oauth_svc.get_valid_access_token(pool, user.id)
    except ml_oauth_svc.MLRefreshError as err:
        raise HTTPException(
            status_code=428,  # Precondition Required — "connect ML first"
            detail=f"ml_oauth_required: {err}",
        )

    # Use the seller_id stored at OAuth-exchange time. Avoids a round-trip to
    # /users/me (which returns 403 on non-Developer-Partner apps) and any
    # fragile parsing of the APP_USR token suffix.
    token_row = await ml_oauth_svc.load_user_tokens(pool, user.id) or {}
    seller_id = token_row.get("ml_user_id")

    try:
        result = await positions_service.check_position(
            item_id=item_id,
            keyword=keyword,
            site_id=site_id,
            category_id=category_id,
            bearer_token=bearer,
            seller_id=seller_id,
        )
    except positions_service.PositionCheckError as err:
        raise HTTPException(status_code=502, detail=f"ml_error: {err}")
    if pool is not None:
        await positions_storage.record_check(
            pool,
            user_id=user.id,
            item_id=result.item_id,
            keyword=result.keyword,
            site_id=result.site_id,
            position=result.position,
            total_results=result.total_results,
            found=result.found,
        )
    return PositionCheckOut(
        itemId=result.item_id,
        keyword=result.keyword,
        position=result.position,
        found=result.found,
        totalResults=result.total_results,
        pagesScanned=result.pages_scanned,
        siteId=result.site_id,
        adsAbove=result.ads_above,
    )


@router.post("/import")
async def import_keywords(
    file: UploadFile = File(..., description="JoomPulse XLSX or CSV (Posição / Palavras-chave / Anúncios)"),
    item_id: str = Query(..., min_length=3, description="MLB to associate keywords with"),
    site_id: str = Query("MLB"),
    limit: int = Query(30, ge=1, le=200, description="Top N keywords by rank to import"),
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    """Bulk import keywords from a JoomPulse export.

    Format expected (Sheet1):
      - row 0: header (Posição na categoria | Palavras-chave | Anúncios)
      - row 1+: rank | keyword | listings_count

    For each row up to `limit`, calls positions_storage.add_tracked
    which is idempotent on (user, item_id, keyword, site_id) — already-
    tracked keywords return existing id without duplication.

    Accepts .xlsx (parsed via openpyxl) and .csv (parsed via stdlib csv).
    """
    if pool is None:
        raise HTTPException(status_code=503, detail="no_db")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".csv")):
        raise HTTPException(status_code=400, detail="file_must_be_xlsx_or_csv")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="empty_file")

    keywords: list[tuple[int | None, str, int | None]] = []  # (rank, keyword, listings)

    if fname.endswith(".xlsx"):
        try:
            from io import BytesIO
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header
                if not row:
                    continue
                rank = None
                keyword: str = ""
                listings = None
                if len(row) >= 1 and row[0] is not None:
                    try:
                        rank = int(float(str(row[0]).strip()))
                    except (ValueError, TypeError):
                        rank = None
                if len(row) >= 2 and row[1] is not None:
                    keyword = str(row[1]).strip()
                if len(row) >= 3 and row[2] is not None:
                    try:
                        listings = int(float(str(row[2]).replace(',', '').replace('.', '').strip()))
                    except (ValueError, TypeError):
                        listings = None
                if keyword and len(keyword) >= 2:
                    keywords.append((rank, keyword, listings))
        except Exception as err:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"xlsx_parse_error: {err}")
    else:
        try:
            import csv
            from io import StringIO
            text = raw.decode("utf-8", errors="replace")
            # Sniff separator: ML Brazil exports often use ; — JoomPulse default is ,
            sep = ";" if text.count(";") > text.count(",") else ","
            reader = csv.reader(StringIO(text, newline=""), delimiter=sep)
            for i, row in enumerate(reader):
                if i == 0:
                    continue
                if not row:
                    continue
                cols = [c.strip().strip('"') for c in row]
                rank = None
                if cols and cols[0]:
                    try:
                        rank = int(float(cols[0]))
                    except (ValueError, TypeError):
                        rank = None
                keyword = cols[1].strip() if len(cols) > 1 else ""
                listings = None
                if len(cols) > 2 and cols[2]:
                    try:
                        listings = int(float(cols[2].replace(',', '').replace('.', '')))
                    except (ValueError, TypeError):
                        listings = None
                if keyword and len(keyword) >= 2:
                    keywords.append((rank, keyword, listings))
        except Exception as err:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"csv_parse_error: {err}")

    if not keywords:
        return {"parsed": 0, "added": 0, "skipped_duplicates": 0, "samples": []}

    # Sort by rank ascending (rank=None goes last); cap to limit
    keywords.sort(key=lambda x: (x[0] if x[0] is not None else 99999))
    keywords = keywords[:limit]

    # Bulk-insert; positions_storage.add_tracked returns (id, was_inserted)
    # so we count net-new vs already-tracked correctly.
    added = 0
    duplicates = 0
    samples: list[dict[str, Any]] = []
    for rank, keyword, listings in keywords:
        try:
            _id, was_inserted = await positions_storage.add_tracked(
                pool,
                user_id=user.id,
                item_id=item_id,
                keyword=keyword,
                site_id=site_id,
                category_id=None,
            )
            if was_inserted:
                added += 1
            else:
                duplicates += 1
            if len(samples) < 5:
                samples.append({"rank": rank, "keyword": keyword, "listings": listings})
        except Exception as err:  # noqa: BLE001
            # Log + continue — one bad row shouldn't kill the import
            samples.append({"rank": rank, "keyword": keyword, "error": str(err)[:100]})

    return {
        "parsed": len(keywords),
        "added": added,
        "skipped_duplicates": duplicates,
        "item_id": item_id,
        "samples": samples,
    }


@router.post("/track", response_model=TrackedKeyword)
async def track(
    body: TrackKeywordIn,
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    if pool is None:
        raise HTTPException(status_code=503, detail="no_db")
    tracked_id, _was_inserted = await positions_storage.add_tracked(
        pool,
        user_id=user.id,
        item_id=body.itemId,
        keyword=body.keyword,
        site_id=body.siteId,
        category_id=body.categoryId,
    )
    return TrackedKeyword(
        id=tracked_id,
        itemId=body.itemId,
        keyword=body.keyword,
        siteId=body.siteId,
        categoryId=body.categoryId,
        createdAt=datetime.now().isoformat(),
        lastPosition=None,
        lastCheckedAt=None,
        lastFound=None,
    )


@router.get("/tracked", response_model=TrackedListOut)
async def tracked(
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    if pool is None:
        return TrackedListOut(tracked=[])
    rows = await positions_storage.list_tracked(pool, user.id)
    return TrackedListOut(
        tracked=[
            TrackedKeyword(
                id=r["id"],
                itemId=r["item_id"],
                keyword=r["keyword"],
                siteId=r["site_id"],
                categoryId=r["category_id"],
                createdAt=_iso(r["created_at"]) or "",
                lastPosition=r["last_position"],
                lastCheckedAt=_iso(r["last_checked_at"]),
                lastFound=r["last_found"],
            )
            for r in rows
        ]
    )


@router.delete("/track/{tracked_id}")
async def untrack(
    tracked_id: int,
    user: CurrentUser = Depends(current_user),
    pool=Depends(get_pool),
):
    if pool is None:
        raise HTTPException(status_code=503, detail="no_db")
    ok = await positions_storage.delete_tracked(pool, user.id, tracked_id)
    if not ok:
        raise HTTPException(status_code=404, detail="not_found")
    return {"deleted": True}
